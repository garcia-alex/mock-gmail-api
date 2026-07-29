from __future__ import annotations

import io
import json

from openpyxl import load_workbook

from mock_gmail_api.docgen import render_xlsx


def _sample_rows() -> list[dict[str, object]]:
    return [
        {"holder": "Jane Doe", "share_class": "Common", "shares": 1_000_000, "ownership_pct": 40.0},
        {
            "holder": "Option Pool",
            "share_class": "Common",
            "shares": 300_000,
            "ownership_pct": 12.0,
        },
    ]


def test_render_produces_valid_xlsx_with_expected_rows() -> None:
    data = render_xlsx.render("Acme Robotics — Cap Table", json.dumps(_sample_rows()))

    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook.active
    assert sheet is not None
    values = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]

    assert values[0][0] == "Acme Robotics — Cap Table"
    assert values[1] == ["Holder", "Share Class", "Shares", "Ownership %"]
    assert values[2] == ["Jane Doe", "Common", 1_000_000, 40.0]


def test_render_strips_markdown_code_fence() -> None:
    fenced = "```json\n" + json.dumps(_sample_rows()) + "\n```"
    data = render_xlsx.render("Title", fenced)

    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook.active
    assert sheet is not None
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) == 4  # title + header + 2 data rows


def test_render_is_deterministic() -> None:
    payload = json.dumps(_sample_rows())
    first = render_xlsx.render("Title", payload)
    second = render_xlsx.render("Title", payload)

    assert first == second
